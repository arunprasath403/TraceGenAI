# tools/excel_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font
from typing import Dict
import os


def generate_excel(testcase_output: Dict, output_path: str) -> str:
    wb = Workbook()

    test_cases = testcase_output["test_cases"]

    # ---------------- Sheet 1: Summary ----------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Total Test Cases"])
    ws_summary.append([len(test_cases)])

    # ---------------- Sheet 2: Detailed Test Cases ----------------
    ws_detail = wb.create_sheet("Detailed Test Cases")
    headers = [
        "Test ID", "Type", "Title", "Preconditions",
        "Steps", "Expected Result",
        "Mapped Requirements", "Mapped APIs", "Mapped UI Elements"
    ]
    ws_detail.append(headers)

    for cell in ws_detail[1]:
        cell.font = Font(bold=True)

    for tc in test_cases:
        ws_detail.append([
            tc["test_id"],
            tc["type"],
            tc["title"],
            "\n".join(tc.get("preconditions", [])),
            "\n".join(tc["steps"]),
            tc["expected_result"],
            ", ".join(tc["mapped_requirements"]),
            ", ".join(tc["mapped_api"]),
            ", ".join(tc["mapped_ui_elements"])
        ])

    # ---------------- Sheet 3: RTM ----------------
    ws_rtm = wb.create_sheet("RTM")
    ws_rtm.append(["Requirement ID", "Mapped Test Cases"])

    rtm_map = {}
    for tc in test_cases:
        for req in tc["mapped_requirements"]:
            rtm_map.setdefault(req, []).append(tc["test_id"])

    for req, tcs in rtm_map.items():
        ws_rtm.append([req, ", ".join(tcs)])

    # ---------------- Sheet 4: API Tests ----------------
    ws_api = wb.create_sheet("API Tests")
    ws_api.append(["API Endpoint", "Test Case ID", "Title"])

    for tc in test_cases:
        for api in tc["mapped_api"]:
            ws_api.append([api, tc["test_id"], tc["title"]])

    # ---------------- Sheet 5: UI Mapping ----------------
    ws_ui = wb.create_sheet("UI Mapping")
    ws_ui.append(["UI Element", "Test Case ID", "Title"])

    for tc in test_cases:
        for ui in tc["mapped_ui_elements"]:
            ws_ui.append([ui, tc["test_id"], tc["title"]])

    # ---------------- Save ----------------
    os.makedirs(output_path, exist_ok=True)
    file_path = os.path.join(output_path, "AI_Generated_Test_Cases.xlsx")
    wb.save(file_path)

    return file_path
